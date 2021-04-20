from openpyxl import Workbook
from openpyxl import load_workbook
import os.path
import time

def caculate_price(file_path):
    # 打开excel文件
    wb = load_workbook(file_path)
    # 选择sheet表
    ws = wb.active
    row_index = 2
    col_index = 1
    data_list = []

    # 创建新生成的excel
    wb_new = Workbook()
    ws_new = wb_new.create_sheet("price",0)

    #取出所有数据
    while ws.cell(row=row_index, column=1).value != None:
        data_dict = {}
        while col_index <= 7:
            data_dict[ws.cell(row=1, column=col_index).value] = ws.cell(row=row_index, column=col_index).value
            col_index += 1
        data_list.append(data_dict)
        col_index = 1
        row_index += 1

    # 新excel写标题
    col_index = 1
    while ws.cell(row=1, column=col_index).value != None:
        ws_new.cell(row=1, column=col_index, value=ws.cell(row=1, column=col_index).value)
        col_index += 1

    total_price = 0
    row_index = 2
    max_list = []
    #筛选出符合的price
    while len(data_list) > 0 :
        for data_dict in data_list.copy():
            col_index = 1

            if data_dict["价税合计"] > 9999:
                data_list.remove(data_dict)
                max_list.append(data_dict)

            if data_dict["价税合计"] + total_price <= 9999:
                total_price = data_dict["价税合计"] + total_price
                data_list.remove(data_dict)
                for key in data_dict.keys():
                    ws_new.cell(row=row_index, column=col_index, value=data_dict[key])
                    col_index += 1
                row_index += 1
        #总计填写
        ws_new.cell(row=row_index, column=6, value=total_price)
        row_index += 1
        total_price = 0
    # 对>9999的数据处理
    for data_dict in max_list:
        col_index = 1
        for key in data_dict.keys():
            ws_new.cell(row=row_index, column=col_index, value=data_dict[key])
            col_index += 1
        row_index += 1
        ws_new.cell(row=row_index, column=6, value="    ")
        row_index += 1
    # 生成新文件
    prev_dir = os.path.dirname(file_path)
    new_path = os.path.join(prev_dir, "{0}.xlsx".format(time.strftime("%Y%m%d%H%M%S", time.localtime())))
    wb_new.save(new_path)
    print("文件生成成功："+new_path.replace('\\\\', '\\'))

if __name__ == '__main__':
    file_path = input("输入文件路径：").replace('\\', '\\\\').replace("\"", "")
    if os.path.exists(file_path):
        caculate_price(file_path)
    else:
        print("文件路径不存在")

    input("Press <enter> end!!!")