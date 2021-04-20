import requests
from lxml import etree
import urllib.parse
import openpyxl

def get_data(search_url, username, password):

    login_url = "http://jira.arcvideo.com:8080/login.jsp"
    select_url = 'http://jira.arcvideo.com:8080/issues/?jql=issuetype%20in%20("Bug%20by%20Client"%2C%20"Bug%20by%20Employee")%20AND%20status%20in%20(Open%2C%20Reopened%2C%20Resolved)%20AND%20text%20~%20' + urllib.parse.quote("\""+search_url+ "\"")
    print(select_url)
    headers = {
        "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.75 Safari/537.36"
    }
    form_data = {
        "os_username": username,
        "os_password": password,
        "os_destination": "",
        "user_role": "",
        "atl_token": "",
        "login": "登录"
    }

    session = requests.session()
    try:
        login_response = session.post(login_url, headers=headers, data=form_data)
        data = login_response.content.decode()
        select_response = session.get(select_url, headers=headers)
        select_data = select_response.content.decode()
        with open("jira.html", "w", encoding="utf-8") as f:
            f.write(select_data)
        return select_data
    except Exception as e:
        print("请求出错：",e)


def parse_data(data):
    x_data = etree.HTML(data)

    issuekey_list = x_data.xpath('//td[@class="issuekey"]/a[@class="issue-link"]/@data-issue-key')
    issuekeyStr_list = format_str(issuekey_list)

    issueType_list = x_data.xpath('//td[@class="issuetype"]//img/@alt')
    issueTypeStr_list = format_str(issueType_list)

    summary_list = x_data.xpath('//td[@class="summary"]//a/text()')
    summaryStr_list = format_str(summary_list)

    projectID_list = x_data.xpath('//td[@class="customfield_10006"]/text()')
    projectIDStr_list = format_str(projectID_list)

    status_list = x_data.xpath('//td[@class="status"]/span/text()')
    statusStr_list = format_str(status_list)

    resolution_list = x_data.xpath('//td[@class="resolution"]/text() | //td[@class="resolution"]/em/text()')
    resolution_list = [x for x in  resolution_list if x !="    " and x!="\n" ]
    resolutionStr_list = format_str(resolution_list)

    issueOrign_list = x_data.xpath('//td[@class="customfield_10001"]/text()')
    issueOrignStr_list = format_str(issueOrign_list)

    issueReason_list = x_data.xpath('//td[@class="customfield_10005"]/text()')
    issueReasonStr_list = format_str(issueReason_list)

    severity_list = x_data.xpath('//td[@class="customfield_10000"]/text()')
    severityStr_list = format_str(severity_list)

    type_list = x_data.xpath('//td[@class="customfield_10004"]/text()')
    typeStr_list = format_str(type_list)

    # pattern = re.compile('//a[@class="issue-link"]/text()')
    # str = pattern.findall(data)
    # print(str)

    priority_list = x_data.xpath('//td[@class="priority"]/img/@alt')
    priorityStr_list = format_str(priority_list)



    num = 0
    num_in = 0
    finalIndex_list = []

    for num in  range(0, len(issuekeyStr_list)):
        finalData_list = []
        finalData_list.append(issueTypeStr_list[num])
        finalData_list.append(priorityStr_list[num])
        finalData_list.append(issuekeyStr_list[num])
        finalData_list.append(summaryStr_list[num])
        finalData_list.append(projectIDStr_list[num])
        finalData_list.append(statusStr_list[num])
        finalData_list.append(resolutionStr_list[num])
        finalData_list.append(issueOrignStr_list[num])
        #finalData_list.append(issueReasonStr_list)值为空，暂时不传
        finalData_list.append(severityStr_list[num])
        finalData_list.append(typeStr_list[num])

        finalIndex_list.append(finalData_list)
        num += 1

    return finalIndex_list

def get_InValue(form_list):
    formFinal_list = []
    for it in form_list:
        if it[7] != "Type 0.Pre-Test":
            formFinal_list.append(it)
        else:
            print("不进行IN值计算：----------{0}".format(it))

    count = 0
    inValue_list = [None]*(len(formFinal_list))
    for count in range(0, len(formFinal_list)):
        sigle_list = []
        sigle_list = formFinal_list[count]
        if sigle_list[9] == "Normal":
            if sigle_list[1] == "high":
                if sigle_list[8] == "1":
                    inValue_list[count] = 30
                elif sigle_list[8] == "2":
                    inValue_list[count] = 15
                elif sigle_list[8] == "3":
                    inValue_list[count] = 8
                elif sigle_list[8] == "4":
                    inValue_list[count] = 3
                elif sigle_list[8] == "5":
                    inValue_list[count] = 2
                elif sigle_list[8] == "6":
                    inValue_list[count] = 0
                else:
                    print("计算失败")

            elif sigle_list[1] == "Medium":
                if sigle_list[8] == "1":
                    inValue_list[count] = 20
                elif sigle_list[8] == "2":
                    inValue_list[count] = 10
                elif sigle_list[8] == "3":
                    inValue_list[count] = 5
                elif sigle_list[8] == "4":
                    inValue_list[count] = 2
                elif sigle_list[8] == "5":
                    inValue_list[count] = 1
                elif sigle_list[8] == "6":
                    inValue_list[count] = 0
                else:
                    print("计算失败")

            elif sigle_list[1] == "Low":
                if sigle_list[8] == "1":
                    inValue_list[count] = 10
                elif sigle_list[8] == "2":
                    inValue_list[count] = 5
                elif sigle_list[8] == "3":
                    inValue_list[count] = 3
                elif sigle_list[8] == "4":
                    inValue_list[count] = 1
                elif sigle_list[8] == "5":
                    inValue_list[count] = 1
                elif sigle_list[8] == "6":
                    inValue_list[count] = 0
                else:
                    print("计算失败")
            else:
                print("计算失败")
        elif sigle_list[9] == "Extreme":
            if sigle_list[1] == "high":
                if sigle_list[8] == "1":
                    inValue_list[count] = 6
                elif sigle_list[8] == "2":
                    inValue_list[count] = 5
                elif sigle_list[8] == "3":
                    inValue_list[count] = 3
                elif sigle_list[8] == "4":
                    inValue_list[count] = 2
                elif sigle_list[8] == "5":
                    inValue_list[count] = 1
                elif sigle_list[8] == "6":
                    inValue_list[count] = 0
                else:
                    print("计算失败")

            elif sigle_list[1] == "Medium":
                if sigle_list[8] == "1":
                    inValue_list[count] = 4
                elif sigle_list[8] == "2":
                    inValue_list[count] = 3
                elif sigle_list[8] == "3":
                    inValue_list[count] = 2
                elif sigle_list[8] == "4":
                    inValue_list[count] = 1
                elif sigle_list[8] == "5":
                    inValue_list[count] = 1
                elif sigle_list[8] == "6":
                    inValue_list[count] = 0
                else:
                    print("计算失败")
            elif sigle_list[1] == "Low":
                if sigle_list[8] == "1":
                    inValue_list[count] = 3
                elif sigle_list[8] == "2":
                    inValue_list[count] = 2
                elif sigle_list[8] == "3":
                    inValue_list[count] = 1
                elif sigle_list[8] == "4":
                    inValue_list[count] = 1
                elif sigle_list[8] == "5":
                    inValue_list[count] = 0
                elif sigle_list[8] == "6":
                    inValue_list[count] = 0
                else:
                    print("计算失败")

            else:
                print("计算失败")
            pass
        else:
            print("计算失败")
        print("IN值{0}".format(inValue_list[count]) + "  ---  {0}".format(sigle_list))

    #总In值计算
    total_value = 0
    for it in inValue_list:
        total_value += it

    return total_value


def format_str(html_list):
    htmlStr_list = []
    for it in html_list:
        str_it = str(it).lstrip().replace("\n", "")
        htmlStr_list.append(str_it)
    return htmlStr_list

# 暂不使用，保存后vba会失效
def xlsm_use(data_require):
    wb = openpyxl.load_workbook("bug IN值统计工具_2018_new.xlsm", read_only=False,keep_vba=True)
    sheet_names = wb.sheetnames
    sheet_caculate = wb[sheet_names[0]]
    row = 2
    for value in data_require:
        sheet_caculate.cell(row, 1, value=value[0])
        sheet_caculate.cell(row, 2, value=value[1])
        sheet_caculate.cell(row, 3, value=value[2])
        sheet_caculate.cell(row, 4, value=value[3])
        sheet_caculate.cell(row, 5, value=value[4])
        sheet_caculate.cell(row, 6, value=value[5])
        sheet_caculate.cell(row, 7, value=value[6])
        sheet_caculate.cell(row, 8, value=value[7])
        sheet_caculate.cell(row, 10, value=value[8])
        sheet_caculate.cell(row, 11, value=value[9])
        row += 1
    wb.save("123.xlsm")  # 注意 excel被手动打开后，保存会失败
    wb.close()


if __name__ == '__main__':
    # username = input("请输入jira用户名：")
    # password = input("请输入密码：")
    search_url = input("请输入项目名称（建议粘贴复制）：")
    html_data = get_data(search_url, "xh3033", "vbZqYmg4")
    index_list = parse_data(html_data)
    inValue = get_InValue(index_list)
    print("IN值总共为：{0}".format(inValue))
