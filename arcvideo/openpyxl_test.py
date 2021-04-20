import openpyxl

wb = openpyxl.load_workbook("bug IN值统计工具_2018_new.xlsm",keep_vba=True)
sheet_names = wb.sheetnames

sheet_caculate = wb[sheet_names[0]]

sheet_caculate.cell(row=2, column=1, value="B1")

wb.save("123.xlsm")  # 注意 excel被手动打开后，保存会失败
wb.close()
